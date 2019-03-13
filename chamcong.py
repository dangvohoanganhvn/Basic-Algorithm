import re
import os
import openpyxl
import pprint
col_date={ '1' :'C' ,
'2' :'D' ,
'3' :'E' ,
'4' :'F' ,
'5' :'G' ,
'6' :'H' ,
'7' :'I' ,
'8' :'J' ,
'9' :'K' ,
'10':'L' ,
'11':'M' ,
'12':'N' ,
'13':'O' ,
'14':'P' , 
'15':'Q' ,
'16':'R' ,
'17':'S' ,
'18':'T' ,
'19':'U' ,
'20':'V' ,
'21':'W' ,
'22':'X' ,
'23':'Y' ,
'24':'Z' ,
'25':'AA' ,
'26':'AB' ,
'27':'AC' ,
'28':'AD' ,
'29':'AE' ,
'30':'AF' ,
'31':'AG'}

row_hour={ '8-9' :'8' ,
'9-10' :'9' ,
'10-11':'10',
'10-12':'11',
'12-13':'12',
'13-14':'13',
'14-15':'14',
'15-16':'15',
'16-17':'16',
'17-18':'17',
'18-19':'18',
'19-20':'19',
'20-21':'20',
'21-22':'21',
'22-23':'22'}

other={ 'haohut':'AA26',
		'giolam':'AH23',
		'thuong':'AA28',
		'phucap':'AA27',
		'tienhaohut':20000,
		}
dir_data=r'C:\Users\HP\Desktop\data_t1.txt'
dir_excel=r'C:\Users\HP\Desktop\toco\Chấm công\luong_t2-draft.xlsx'
workhour=re.compile(r'''
([a-zA-Z]+) #name
-(\d{1,2}) #ngay
(=|:) #=symbol
((\d{1,2})-(\d{1,2})|off)
''',re.VERBOSE)

def get_data(rule,dir_data):
	file=open(dir_data,'r')
	raw_data=file.read()
	data=rule.findall(raw_data)
	return data

#wb=openpyxl.load_workbook(dir_excel)
# 
#sheet=wb.get_sheet_names()
#
#data=get_data(workhour,dir_data) # data of staff
#print(data)
#sheet=wb.get_sheet_by_name(data[1][0].lower())
#print(sheet.title)
#print( sheet[cell_date[data[1][1]]+'12'].value)
#sheet[cell_date[data[1][1]]+'12'].value='v'
#print( sheet[cell_date[data[1][1]]+'12'].value)
#wb.save(dir_excel)
def stick_data(sheet,date,begin,end,workhour):
	print ('stick for '+ sheet.title+' ngay '+ date +' ' +workhour)
	for i in range(int(begin),int(end)):
		sheet[col_date[date]+str(i)]='v'


def populate_data(excel,raw_data):
	wb=openpyxl.load_workbook(excel)
	data=get_data(workhour,raw_data) # data of staff
	pprint.pprint(data)
	for staff in data: 
		sheet=wb.get_sheet_by_name(staff[0].lower())
		if str(staff[3])!="off":
			stick_data(sheet,staff[1],staff[4],staff[5],staff[3])
		wb.save(excel)


def hao_hut(money,pos,excel):
	wb=openpyxl.load_workbook(excel)
	sheet_name=wb.get_sheet_names()
	print (sheet_name)	 
	for staff in sheet_name:
		print ('hao hut for '+str(staff) +' ' + str(money))
		sheet=wb.get_sheet_by_name(staff)
		sheet[str(pos)]=money
	wb.save(excel)
	
def thuong(total_hours,thuong,excel):
	wb=openpyxl.load_workbook(excel,data_only=True)
	chamcong=openpyxl.load_workbook(excel)
	sheet_name=wb.get_sheet_names()
	#print (sheet_name)	
	
	for staff in sheet_name:	
		sheet=wb.get_sheet_by_name(staff)
		
		if sheet[total_hours].value>=200:
			chamcong.get_sheet_by_name(staff)[thuong]=200000
			print ('thuong 200000 for '+str(staff))
		elif sheet[total_hours].value>=150:
			chamcong.get_sheet_by_name(staff)[thuong]=150000
			print ('thuong 150000 for '+str(staff))
		elif sheet[total_hours].value>=100:
			chamcong.get_sheet_by_name(staff)[thuong]=100000
			print ('thuong 100000 for '+str(staff))
		if staff=='nhu':
			chamcong.get_sheet_by_name(staff)['AA27']=600000
	chamcong.save(excel)
#main

#populate_data(dir_excel,dir_data)
#hao_hut(other['tienhaohut'],other['haohut'],dir_excel)
thuong(other['giolam'],other['thuong'],dir_excel)
print (dir_data)
print  (dir_excel)

	